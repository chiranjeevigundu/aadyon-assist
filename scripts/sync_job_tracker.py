#!/usr/bin/env python3
"""Sync a job-tracker spreadsheet into /api/applications (idempotent upsert).

Reads a `.xlsx` job tracker (the kind the daily job-search digest updates),
maps its columns to the `applications` entity by fuzzy header matching, and
upserts each row into a running API: create rows that don't exist yet, PATCH
only the fields that actually changed, and leave unchanged rows untouched.

The Digital Me "Career" dimension reads straight from `applications`, so a
successful sync is immediately reflected there.

Usage:
    # dry run — auth + read the live rows, print the plan, write nothing
    python scripts/sync_job_tracker.py --xlsx JOB_TRACKER.xlsx --dry-run \
        --base http://localhost:8000 --email you@example.com --password ...

    # for real
    python scripts/sync_job_tracker.py --xlsx JOB_TRACKER.xlsx \
        --base http://localhost:8000 --token "$AADYON_TOKEN"

Auth (first that resolves wins): --token / $AADYON_TOKEN, else login with
--email/--password (or $AADYON_EMAIL / $AADYON_PASSWORD). Never hard-code
secrets — pass a token or credentials via env (see `just sync-jobs`).

Design notes (the parts that are easy to get wrong):
  - Matching: an xlsx row maps to an existing application by the natural key
    (company, role), compared case-insensitively.
  - Equality: values from the xlsx and values echoed back by the API are both
    reduced to one canonical form before comparison — dates to 'YYYY-MM-DD',
    money to a 2-decimal float (Postgres numeric(12,2), which the API returns
    as a JSON float), text stripped. Without this, "2026-07-15" != a datetime
    and 120000 != 120000.0 would PATCH every row on every run.
  - Blanks never clear data: an empty xlsx cell is dropped from the payload, so
    a sparse tracker can't wipe fields the API already holds.

This is an ops script (like scripts/verify.py); openpyxl is imported lazily so
the pure-logic helpers stay unit-testable without it installed.
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys

import requests

# Writable columns of the `applications` entity (app/models/tables.py). Order
# here is only cosmetic (drives --verbose output).
FIELDS = [
    "company", "role", "status", "salary_min", "salary_max", "location",
    "work_type", "source", "url", "applied_date", "notes",
]
FLOAT_FIELDS = {"salary_min", "salary_max"}
DATE_FIELDS = {"applied_date"}
KEY_FIELDS = ("company", "role")  # natural key for matching xlsx <-> API rows

# Known application stages (applications.status). Anything else is passed through
# lowercased — the column is free text, so we don't reject unknown stages.
STATUSES = {"saved", "applied", "screening", "interview", "offer", "rejected", "accepted"}

# Fuzzy header aliases: a spreadsheet header is normalised (lowercased, stripped
# of non-alphanumerics) and matched against these. The field name itself always
# matches. First column to claim a field wins.
ALIASES: dict[str, list[str]] = {
    "company": ["company", "employer", "organization", "organisation", "org", "companyname"],
    "role": ["role", "title", "position", "jobtitle", "job", "roletitle"],
    "status": ["status", "stage", "state", "progress"],
    "salary_min": ["salarymin", "minsalary", "salaryfrom", "salarylow", "minpay", "compmin",
                   "salary", "pay", "compensation", "comp", "ctc"],
    "salary_max": ["salarymax", "maxsalary", "salaryto", "salaryhigh", "maxpay", "compmax"],
    "location": ["location", "city", "place", "where", "geo"],
    "work_type": ["worktype", "arrangement", "modality", "workmode", "remotehybrid", "remoteonsite",
                  "remote", "onsite"],
    "source": ["source", "via", "channel", "board", "jobboard", "referral"],
    "url": ["url", "link", "joburl", "posting", "postingurl", "jobposting", "jobling", "listing"],
    "applied_date": ["applieddate", "dateapplied", "applied", "appliedon", "date", "dateaded",
                     "dateadded"],
    "notes": ["notes", "note", "comments", "comment", "remarks", "details"],
}

_DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%b-%Y", "%b %d, %Y")


# --------------------------------------------------------------------------- value canon
def _norm_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text).strip().lower())


def to_float(value) -> float | None:
    """Parse money-ish cells to a float. Handles $, commas, and a trailing k."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):  # bool is an int subclass — never a salary
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().lower().replace("$", "").replace(",", "").replace(" ", "")
    if not s:
        return None
    mult = 1.0
    if s.endswith("k"):
        mult, s = 1000.0, s[:-1]
    try:
        return float(s) * mult
    except ValueError:
        return None


def to_iso_date(value) -> str | None:
    """Reduce a date cell (datetime, date, or string) to 'YYYY-MM-DD'."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    s = str(value).strip()
    if not s:
        return None
    # ISO with a time component, e.g. what the API might echo: '2026-07-15T00:00:00'
    head = s.split("T", 1)[0].split(" ", 1)[0]
    for fmt in _DATE_FORMATS:
        try:
            return dt.datetime.strptime(head if fmt == "%Y-%m-%d" else s, fmt).date().isoformat()
        except ValueError:
            continue
    # Last resort: a bare ISO-looking head
    try:
        return dt.date.fromisoformat(head).isoformat()
    except ValueError:
        return None


def canon(field: str, value):
    """Canonical, comparable + API-ready form of a field value.

    Applied to both xlsx cells and API responses so equality is meaningful.
    Returns None for empty values (callers drop None so blanks never overwrite).
    """
    if value is None:
        return None
    if field in FLOAT_FIELDS:
        f = to_float(value)
        return round(f, 2) if f is not None else None
    if field in DATE_FIELDS:
        return to_iso_date(value)
    s = str(value).strip()
    if field == "status":
        s = s.lower()
    return s or None


# --------------------------------------------------------------------------- xlsx read
def map_headers(headers: list) -> dict[int, str]:
    """Map column index -> field name via fuzzy header matching (first wins)."""
    used: set[str] = set()
    mapping: dict[int, str] = {}
    for idx, raw in enumerate(headers):
        if raw is None:
            continue
        norm = _norm_header(raw)
        if not norm:
            continue
        for field in FIELDS:
            if field in used:
                continue
            if norm == field or norm in ALIASES.get(field, []):
                mapping[idx] = field
                used.add(field)
                break
    return mapping


def read_tracker(path: str, sheet: str | None = None) -> list[dict]:
    """Read the tracker into a list of desired records ({field: canon value}).

    Rows without a company are skipped (company is NOT NULL in the schema).
    """
    from openpyxl import load_workbook  # lazy: keeps helpers importable without the dep

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows_iter = ws.iter_rows(values_only=True)

    headers = None
    for row in rows_iter:
        if row and any(c is not None and str(c).strip() for c in row):
            headers = list(row)
            break
    if headers is None:
        return []
    mapping = map_headers(headers)
    if "company" not in mapping.values():
        raise SystemExit(
            "Could not find a 'company' column in the tracker. Headers seen: "
            + ", ".join(str(h) for h in headers if h is not None)
        )

    records: list[dict] = []
    for row in rows_iter:
        if not row or not any(c is not None and str(c).strip() for c in row):
            continue
        rec: dict = {}
        for idx, field in mapping.items():
            if idx < len(row):
                v = canon(field, row[idx])
                if v is not None:
                    rec[field] = v
        if rec.get("company"):
            records.append(rec)
    return records


# --------------------------------------------------------------------------- diff
def _key(rec: dict) -> tuple:
    # Match key only — case-folded so "Acme Corp" and "acme corp" collide.
    # The stored/created value keeps its original case (canon doesn't lowercase).
    return tuple(v.lower() if isinstance(v := rec.get(f), str) else v for f in KEY_FIELDS)


def index_existing(api_rows: list[dict]) -> dict[tuple, dict]:
    """Index API rows by natural key, canonicalised. First row per key wins."""
    idx: dict[tuple, dict] = {}
    for row in api_rows:
        c = {f: canon(f, row.get(f)) for f in FIELDS}
        c["id"] = row.get("id")
        idx.setdefault(_key(c), c)
    return idx


def diff(desired: dict, existing: dict | None) -> dict:
    """Fields in `desired` whose canonical value differs from `existing`.

    `existing` already-canonical (from index_existing) or None for a new row.
    Only present desired fields are considered, so blanks never clear data.
    """
    changed = {}
    for field, want in desired.items():
        if field == "id":
            continue
        if existing is None or existing.get(field) != want:
            changed[field] = want
    return changed


def plan(desired_rows: list[dict], existing_index: dict[tuple, dict]) -> tuple[list, list, list]:
    """Return (creates, updates, unchanged). updates carry (id, changed-fields)."""
    creates, updates, unchanged = [], [], []
    for rec in desired_rows:
        match = existing_index.get(_key(rec))
        if match is None:
            creates.append(rec)
        else:
            changed = diff(rec, match)
            (updates if changed else unchanged).append(
                (match["id"], rec, changed) if changed else (match["id"], rec, {})
            )
    return creates, updates, unchanged


# --------------------------------------------------------------------------- api
class Api:
    def __init__(self, base: str, token: str, timeout: int = 30):
        self.base = base.rstrip("/")
        self.timeout = timeout
        self.s = requests.Session()
        self.s.headers["Authorization"] = f"Bearer {token}"

    def list_applications(self) -> list[dict]:
        r = self.s.get(f"{self.base}/api/applications", timeout=self.timeout)
        r.raise_for_status()
        return r.json()

    def create(self, payload: dict) -> dict:
        r = self.s.post(f"{self.base}/api/applications", json=payload, timeout=self.timeout)
        r.raise_for_status()
        return r.json()

    def update(self, row_id: str, payload: dict) -> dict:
        r = self.s.patch(f"{self.base}/api/applications/{row_id}", json=payload, timeout=self.timeout)
        r.raise_for_status()
        return r.json()


def resolve_token(args) -> str:
    token = args.token or os.environ.get("AADYON_TOKEN")
    if token:
        return token.strip()
    email = args.email or os.environ.get("AADYON_EMAIL")
    password = args.password or os.environ.get("AADYON_PASSWORD")
    if not (email and password):
        raise SystemExit(
            "No credentials: pass --token/$AADYON_TOKEN, or --email/--password "
            "(or $AADYON_EMAIL / $AADYON_PASSWORD)."
        )
    r = requests.post(
        f"{args.base.rstrip('/')}/api/auth/login",
        json={"email": email, "password": password}, timeout=30,
    )
    if r.status_code != 200:
        raise SystemExit(f"Login failed ({r.status_code}): {r.text[:200]}")
    return r.json()["token"]


def _fmt(rec: dict) -> str:
    return f"{rec.get('company', '?')} — {rec.get('role', '?')}"


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Sync a job-tracker xlsx into /api/applications")
    ap.add_argument("--xlsx", required=True, help="path to the job tracker .xlsx")
    ap.add_argument("--sheet", default=None, help="sheet name (default: active/first sheet)")
    ap.add_argument("--base", default=os.environ.get("AADYON_BASE", "http://localhost:8000"),
                    help="base URL of a running API")
    ap.add_argument("--token", default=None, help="JWT bearer token (or $AADYON_TOKEN)")
    ap.add_argument("--email", default=None, help="login email (or $AADYON_EMAIL)")
    ap.add_argument("--password", default=None, help="login password (or $AADYON_PASSWORD)")
    ap.add_argument("--dry-run", action="store_true", help="read + plan only; write nothing")
    ap.add_argument("--verbose", action="store_true", help="print every field change")
    args = ap.parse_args(argv)

    desired = read_tracker(args.xlsx, args.sheet)
    print(f"Read {len(desired)} row(s) from {args.xlsx}")

    token = resolve_token(args)
    api = Api(args.base, token)
    try:
        existing_rows = api.list_applications()
    except requests.RequestException as e:
        raise SystemExit(f"Could not reach {args.base}: {e}") from e
    existing_index = index_existing(existing_rows)
    print(f"API currently has {len(existing_rows)} application(s)")

    creates, updates, unchanged = plan(desired, existing_index)
    print(f"Plan: +{len(creates)} create, ~{len(updates)} update, ={len(unchanged)} unchanged")

    for rec in creates:
        print(f"  + {_fmt(rec)}")
    for _id, rec, changed in updates:
        print(f"  ~ {_fmt(rec)}: {', '.join(sorted(changed))}")
        if args.verbose:
            for k in sorted(changed):
                print(f"      {k} -> {changed[k]!r}")

    if args.dry_run:
        print("Dry run — no changes written.")
        return 0

    created = updated = 0
    try:
        for rec in creates:
            api.create(rec)
            created += 1
        for _id, _rec, changed in updates:
            api.update(_id, changed)
            updated += 1
    except requests.RequestException as e:
        body = getattr(e.response, "text", "")[:200] if getattr(e, "response", None) else ""
        raise SystemExit(f"Write failed after {created} create/{updated} update: {e} {body}") from e

    print(f"Done: {created} created, {updated} updated, {len(unchanged)} unchanged.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
